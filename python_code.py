import serial
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

# Specify the path where you want to save the Excel file
excel_file_path = "SensorData.xlsx"

# Specify the sheet name where you want to save the data
sheet_name = "DataSheet"

# Open the serial port (replace 'COM19' with the actual COM port)
ser = serial.Serial('COM19', 115200)  # Replace 'COM19' with the actual COM port and adjust baudrate if needed

# Load the existing Excel workbook
try:
    workbook = load_workbook(excel_file_path)
    # If the sheet already exists, load it; otherwise, create a new one
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)
except FileNotFoundError:
    # If the Excel file doesn't exist, create a new one
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

# Set the headers for the sensor data, including the timestamp
sheet["A1"] = "Timestamp"
sheet["B1"] = "Temperature (°C)"
sheet["C1"] = "Pressure (BAR)"
sheet["D1"] = "Flow Rate"

# Find the next row to write data
next_row = sheet.max_row + 1

try:
    while True:
        # Read data from the serial port
        data = ser.readline().decode().strip()
        print(data)  # Print the incoming data
        
        # Split the data into individual values using a comma as the delimiter
        values = data.split(",")

        # Check if the data contains the expected number of values
        if len(values) == 3:  # Adjust according to the number of values received
            temperature, pressure, flow_rate = map(float, values)

            # Get the current timestamp
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Append the sensor values and timestamp to the Excel sheet in the next row
            row = [timestamp, temperature, pressure, flow_rate]
            sheet.append(row)

            # Save the workbook with sensor data to an Excel file
            workbook.save(excel_file_path)
            print(f"Saved data to {excel_file_path} in row {next_row}")

            # Increment the next row for the next data
            next_row += 1

        else:
            print(f"Received invalid data: {data}")

except KeyboardInterrupt:
    # Handle a Ctrl+C interrupt by gracefully closing the serial port and saving the Excel file
    ser.close()
    workbook.save(excel_file_path)
    print("Script interrupted. Excel file saved.")
